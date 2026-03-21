"""
Export Census Data - CSV/Power BI/Excel Export

This management command exports city demographic data for visualization and analysis.

Usage:
    python manage.py export_census_data --format csv
    python manage.py export_census_data --format csv --site seekingspringfield.com
    python manage.py export_census_data --format excel --output springfield_data.xlsx
    python manage.py export_census_data --format json
    python manage.py export_census_data --format powerbi  # Power BI optimized CSV

Export all 24 Springfields for comparison:
    python manage.py export_census_data --site seekingspringfield.com --format powerbi

Course Module: youbetyourazure.com/courses/power-bi-visualization
"""

from django.core.management.base import BaseCommand
from django.utils import timezone
from core.models import Site, City, Article, Business
import csv
import json
import os
from datetime import datetime


class Command(BaseCommand):
    help = 'Export census and city data for Power BI/Excel/CSV'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--format',
            type=str,
            default='csv',
            choices=['csv', 'excel', 'json', 'powerbi'],
            help='Export format (default: csv)'
        )
        
        parser.add_argument(
            '--site',
            type=str,
            help='Filter by site domain (e.g., seekingspringfield.com)'
        )
        
        parser.add_argument(
            '--output',
            type=str,
            help='Output filename (default: auto-generated)'
        )
        
        parser.add_argument(
            '--include-articles',
            action='store_true',
            help='Include article count and performance data'
        )
        
        parser.add_argument(
            '--include-businesses',
            action='store_true',
            help='Include business count and revenue data'
        )
    
    def handle(self, *args, **options):
        """Main execution"""
        
        self.stdout.write("=" * 80)
        self.stdout.write(self.style.SUCCESS("CENSUS DATA EXPORT"))
        self.stdout.write("=" * 80)
        self.stdout.write("")
        
        # Get cities to export
        cities = self._get_cities(options)
        
        if not cities.exists():
            self.stdout.write(self.style.WARNING("⚠️  No cities found matching criteria"))
            return
        
        export_format = options['format']
        output_file = options.get('output') or self._generate_filename(options)
        
        self.stdout.write(f"📊 EXPORT PLAN:")
        self.stdout.write(f"   Cities: {cities.count()}")
        self.stdout.write(f"   Format: {export_format.upper()}")
        self.stdout.write(f"   Output: {output_file}")
        if options.get('site'):
            self.stdout.write(f"   Site Filter: {options['site']}")
        self.stdout.write("")
        
        # Export based on format
        if export_format == 'csv':
            self._export_csv(cities, output_file, options)
        elif export_format == 'powerbi':
            self._export_powerbi(cities, output_file, options)
        elif export_format == 'json':
            self._export_json(cities, output_file, options)
        elif export_format == 'excel':
            self._export_excel(cities, output_file, options)
        
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"✅ Export complete: {output_file}"))
        self.stdout.write("")
        
        # Power BI instructions
        if export_format == 'powerbi':
            self._show_powerbi_instructions(output_file)
    
    def _get_cities(self, options):
        """Build queryset of cities to export"""
        
        queryset = City.objects.all()
        
        # Filter by site
        if options.get('site'):
            site_domain = options['site']
            queryset = queryset.filter(sites__domain_name=site_domain)
        
        # Only export cities with Census data
        queryset = queryset.filter(population__isnull=False)
        
        return queryset.distinct().order_by('state', 'name')
    
    def _generate_filename(self, options):
        """Generate output filename"""
        
        timestamp = datetime.now().strftime('%Y%m%d')
        export_format = options['format']
        
        if options.get('site'):
            site = options['site'].replace('.com', '').replace('.', '_')
            base = f"{site}_{timestamp}"
        else:
            base = f"census_export_{timestamp}"
        
        if export_format == 'excel':
            return f"{base}.xlsx"
        elif export_format == 'json':
            return f"{base}.json"
        else:
            return f"{base}.csv"
    
    def _export_csv(self, cities, output_file, options):
        """Export to CSV format"""
        
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = self._get_csv_fieldnames(options)
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for city in cities:
                row = self._build_city_row(city, options)
                writer.writerow(row)
        
        self.stdout.write(f"   💾 Saved {cities.count()} cities to CSV")
    
    def _export_powerbi(self, cities, output_file, options):
        """
        Export Power BI optimized CSV
        
        Power BI formatting:
        - Clean column names (no underscores)
        - Proper data types (remove $ and % symbols, Power BI will format)
        - Date/time in ISO format
        - Boolean as TRUE/FALSE
        """
        
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = self._get_powerbi_fieldnames(options)
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for city in cities:
                row = self._build_powerbi_row(city, options)
                writer.writerow(row)
        
        self.stdout.write(f"   💾 Saved {cities.count()} cities to Power BI CSV")
    
    def _export_json(self, cities, output_file, options):
        """Export to JSON format"""
        
        data = {
            'export_date': timezone.now().isoformat(),
            'city_count': cities.count(),
            'cities': []
        }
        
        for city in cities:
            city_data = {
                'name': city.name,
                'state': city.state,
                'country': city.country,
                'population': city.population,
                'median_income': float(city.median_income) if city.median_income else None,
                'census_data': city.census_data or {},
                'last_updated': city.last_census_updated.isoformat() if city.last_census_updated else None,
            }
            
            if options.get('include_articles'):
                city_data['article_count'] = city.article_set.count()
            
            if options.get('include_businesses'):
                city_data['business_count'] = city.business_set.count()
            
            data['cities'].append(city_data)
        
        with open(output_file, 'w', encoding='utf-8') as jsonfile:
            json.dump(data, jsonfile, indent=2, default=str)
        
        self.stdout.write(f"   💾 Saved {cities.count()} cities to JSON")
    
    def _export_excel(self, cities, output_file, options):
        """Export to Excel format (requires openpyxl)"""
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Census Data"
            
            # Header row with formatting
            fieldnames = self._get_csv_fieldnames(options)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, field in enumerate(fieldnames, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = field.replace('_', ' ').title()
                cell.fill = header_fill
                cell.font = header_font
            
            # Data rows
            for row_num, city in enumerate(cities, 2):
                row_data = self._build_city_row(city, options)
                for col_num, field in enumerate(fieldnames, 1):
                    sheet.cell(row=row_num, column=col_num).value = row_data.get(field)
            
            # Auto-size columns
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            workbook.save(output_file)
            self.stdout.write(f"   💾 Saved {cities.count()} cities to Excel")
        
        except ImportError:
            self.stdout.write(self.style.ERROR("❌ openpyxl not installed"))
            self.stdout.write("Install with: pip install openpyxl")
            self.stdout.write("Falling back to CSV...")
            self._export_csv(cities, output_file.replace('.xlsx', '.csv'), options)
    
    def _get_csv_fieldnames(self, options):
        """Get CSV column names"""
        
        fields = [
            'city_name',
            'state',
            'country',
            'population',
            'median_income',
            'median_home_value',
            'median_age',
            'unemployment_rate',
            'college_education_rate',
            'homeownership_rate',
            'last_updated',
        ]
        
        if options.get('include_articles'):
            fields.extend(['article_count', 'published_article_count'])
        
        if options.get('include_businesses'):
            fields.extend(['business_count', 'premium_business_count', 'monthly_revenue'])
        
        return fields
    
    def _get_powerbi_fieldnames(self, options):
        """Get Power BI optimized column names (no underscores)"""
        
        fields = [
            'City',
            'State',
            'Country',
            'Population',
            'Median Income',
            'Median Home Value',
            'Median Age',
            'Unemployment Rate',
            'College Education Rate',
            'Homeownership Rate',
            'Last Updated',
        ]
        
        if options.get('include_articles'):
            fields.extend(['Article Count', 'Published Articles'])
        
        if options.get('include_businesses'):
            fields.extend(['Business Count', 'Premium Businesses', 'Monthly Revenue'])
        
        return fields
    
    def _build_city_row(self, city, options):
        """Build data row for a city"""
        
        census = city.census_data or {}
        
        row = {
            'city_name': city.name,
            'state': city.state,
            'country': city.country,
            'population': city.population or '',
            'median_income': f"${city.median_income:.0f}" if city.median_income else '',
            'median_home_value': f"${census.get('median_home_value', 0):.0f}" if census.get('median_home_value') else '',
            'median_age': f"{census.get('median_age', 0):.1f}" if census.get('median_age') else '',
            'unemployment_rate': f"{census.get('unemployment_rate', 0):.1f}%" if census.get('unemployment_rate') else '',
            'college_education_rate': f"{census.get('college_education_rate', 0):.1f}%" if census.get('college_education_rate') else '',
            'homeownership_rate': f"{census.get('homeownership_rate', 0):.1f}%" if census.get('homeownership_rate') else '',
            'last_updated': city.last_census_updated.strftime('%Y-%m-%d') if city.last_census_updated else '',
        }
        
        if options.get('include_articles'):
            row['article_count'] = city.article_set.count()
            row['published_article_count'] = city.article_set.filter(published=True).count()
        
        if options.get('include_businesses'):
            businesses = city.business_set.filter(active=True)
            row['business_count'] = businesses.count()
            row['premium_business_count'] = businesses.filter(tier__in=['PREMIUM', 'FEATURED']).count()
            row['monthly_revenue'] = sum(b.monthly_cost for b in businesses if b.monthly_cost)
        
        return row
    
    def _build_powerbi_row(self, city, options):
        """Build Power BI optimized row (clean numbers, no formatting)"""
        
        census = city.census_data or {}
        
        row = {
            'City': city.name,
            'State': city.state,
            'Country': city.country,
            'Population': city.population or 0,
            'Median Income': float(city.median_income) if city.median_income else 0,
            'Median Home Value': float(census.get('median_home_value', 0)),
            'Median Age': float(census.get('median_age', 0)),
            'Unemployment Rate': float(census.get('unemployment_rate', 0)),
            'College Education Rate': float(census.get('college_education_rate', 0)),
            'Homeownership Rate': float(census.get('homeownership_rate', 0)),
            'Last Updated': city.last_census_updated.isoformat() if city.last_census_updated else '',
        }
        
        if options.get('include_articles'):
            row['Article Count'] = city.article_set.count()
            row['Published Articles'] = city.article_set.filter(published=True).count()
        
        if options.get('include_businesses'):
            businesses = city.business_set.filter(active=True)
            row['Business Count'] = businesses.count()
            row['Premium Businesses'] = businesses.filter(tier__in=['PREMIUM', 'FEATURED']).count()
            row['Monthly Revenue'] = float(sum(b.monthly_cost for b in businesses if b.monthly_cost))
        
        return row
    
    def _show_powerbi_instructions(self, filename):
        """Show instructions for importing into Power BI"""
        
        self.stdout.write("=" * 80)
        self.stdout.write(self.style.SUCCESS("📊 POWER BI IMPORT INSTRUCTIONS"))
        self.stdout.write("=" * 80)
        self.stdout.write("")
        self.stdout.write("STEP 1: Open Power BI Desktop")
        self.stdout.write("")
        self.stdout.write("STEP 2: Get Data > Text/CSV")
        self.stdout.write(f"   Select file: {filename}")
        self.stdout.write("")
        self.stdout.write("STEP 3: Transform Data")
        self.stdout.write("   - Change data types if needed")
        self.stdout.write("   - Population → Whole Number")
        self.stdout.write("   - Median Income → Currency")
        self.stdout.write("   - Rates → Percentage")
        self.stdout.write("")
        self.stdout.write("STEP 4: Create Visualizations")
        self.stdout.write("   📈 SUGGESTED VISUALS:")
        self.stdout.write("   - Map: Population by City (bubble size)")
        self.stdout.write("   - Bar Chart: Top 10 Cities by Income")
        self.stdout.write("   - Table: All Demographics (sortable)")
        self.stdout.write("   - Card: Total Population Across All Cities")
        self.stdout.write("   - Scatter Plot: Income vs Home Value")
        self.stdout.write("")
        self.stdout.write("STEP 5: Add Slicers")
        self.stdout.write("   - State (filter cities by state)")
        self.stdout.write("   - Population Range")
        self.stdout.write("")
        self.stdout.write("COURSE: youbetyourazure.com/courses/power-bi-visualization")
        self.stdout.write("Learn advanced Power BI + Django integration!")
        self.stdout.write("")
        self.stdout.write("=" * 80)
